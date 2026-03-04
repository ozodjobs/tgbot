import io

import numpy as np
from PIL import Image, ImageFilter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.utils import ImageReader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# OpenCV for document edge detection
try:
    import cv2
    HAS_CV2 = True
except ImportError:
    HAS_CV2 = False

# scipy for morphological ops (optional boost)
try:
    from scipy import ndimage as ndi
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False


# ================================================================
# ══════════  AUTO-CROP DOCUMENT EDGES (CamScanner style)  ═══════
# ================================================================

def order_points(pts: np.ndarray) -> np.ndarray:
    rect = np.zeros((4, 2), dtype=np.float32)
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]
    rect[2] = pts[np.argmax(s)]
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]
    rect[3] = pts[np.argmax(diff)]
    return rect


def four_point_transform(image: np.ndarray, pts: np.ndarray) -> np.ndarray:
    rect = order_points(pts)
    (tl, tr, br, bl) = rect
    widthA  = np.linalg.norm(br - bl)
    widthB  = np.linalg.norm(tr - tl)
    maxW    = max(int(widthA), int(widthB))
    heightA = np.linalg.norm(tr - br)
    heightB = np.linalg.norm(tl - bl)
    maxH    = max(int(heightA), int(heightB))
    dst = np.array([
        [0, 0],
        [maxW - 1, 0],
        [maxW - 1, maxH - 1],
        [0, maxH - 1]
    ], dtype=np.float32)
    M       = cv2.getPerspectiveTransform(rect, dst)
    warped  = cv2.warpPerspective(image, M, (maxW, maxH))
    return warped


def auto_crop_document(pil_img: Image.Image) -> Image.Image:
    if not HAS_CV2:
        return pil_img
    orig = np.array(pil_img.convert("RGB"))
    h, w = orig.shape[:2]
    scale     = 800 / max(h, w)
    small     = cv2.resize(orig, (int(w * scale), int(h * scale)))
    gray      = cv2.cvtColor(small, cv2.COLOR_RGB2GRAY)
    blurred   = cv2.bilateralFilter(gray, 9, 75, 75)
    med       = float(np.median(blurred))
    lo, hi    = int(max(0, 0.66 * med)), int(min(255, 1.33 * med))
    edges     = cv2.Canny(blurred, lo, hi)
    edges     = cv2.dilate(edges, np.ones((3, 3), np.uint8), iterations=1)
    cnts, _   = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    cnts      = sorted(cnts, key=cv2.contourArea, reverse=True)[:10]
    screen_cnt = None
    for c in cnts:
        peri   = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)
        if len(approx) == 4:
            area = cv2.contourArea(approx)
            if area > 0.15 * small.shape[0] * small.shape[1]:
                screen_cnt = approx
                break
    if screen_cnt is None:
        return pil_img
    pts      = screen_cnt.reshape(4, 2).astype(np.float32)
    pts     /= scale
    warped   = four_point_transform(orig, pts)
    return Image.fromarray(warped)


# ================================================================
# ══════════  PROFESSIONAL SCAN PROCESSING PIPELINE  ═════════════
# ================================================================

def process_scan_image(img: Image.Image) -> Image.Image:
    img = auto_crop_document(img)
    img = img.convert("L")
    orig_w, orig_h = img.size
    long_side = max(orig_w, orig_h)
    if long_side < 2400:
        factor = 2400 / long_side
        img = img.resize(
            (int(orig_w * factor), int(orig_h * factor)),
            Image.LANCZOS
        )
    arr = np.array(img, dtype=np.uint8)
    if HAS_CV2:
        arr    = cv2.fastNlMeansDenoising(arr, h=10, templateWindowSize=7, searchWindowSize=21)
        clahe  = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        arr    = clahe.apply(arr)
        binary = cv2.adaptiveThreshold(arr, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                       cv2.THRESH_BINARY, blockSize=51, C=18)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
        binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN,  kernel)
        blurred_sharp = cv2.GaussianBlur(binary, (0, 0), 1.2)
        binary        = cv2.addWeighted(binary, 1.8, blurred_sharp, -0.8, 0)
        _, binary     = cv2.threshold(binary, 127, 255, cv2.THRESH_BINARY)
        result_arr = binary
    else:
        blurred = Image.fromarray(arr).filter(ImageFilter.GaussianBlur(radius=1))
        arr     = np.array(blurred, dtype=np.float32)
        p2, p98 = np.percentile(arr, 2), np.percentile(arr, 98)
        if p98 > p2:
            arr = np.clip((arr - p2) / (p98 - p2) * 255, 0, 255)
        arr_u8  = arr.astype(np.uint8)
        bg_img  = Image.fromarray(arr_u8).filter(ImageFilter.GaussianBlur(radius=40))
        bg      = np.array(bg_img, dtype=np.float32)
        diff    = bg.astype(np.float32) - arr.astype(np.float32)
        binary  = np.where(diff > 20, 0, 255).astype(np.uint8)
        binary  = np.where(binary < 128, 0, 255).astype(np.uint8)
        sharp_img = Image.fromarray(binary)
        sharp_img = sharp_img.filter(ImageFilter.UnsharpMask(radius=1, percent=200, threshold=2))
        binary    = np.where(np.array(sharp_img) < 128, 0, 255).astype(np.uint8)
        result_arr = binary
    result_arr = np.where(result_arr >= 220, 255,
                 np.where(result_arr <= 80,    0, result_arr)).astype(np.uint8)
    result_img = Image.fromarray(result_arr, mode="L")
    out = Image.new("RGB", result_img.size, (255, 255, 255))
    out.paste(result_img)
    return out


# ================================================================
# ═════════════  PDF BUILDER  ════════════════════════════════════
# ================================================================

def build_pdf_with_reportlab(pil_images: list, pdf_path: str):
    PAGE_W, PAGE_H = A4
    MARGIN   = 20
    usable_w = PAGE_W - 2 * MARGIN
    usable_h = PAGE_H - 2 * MARGIN
    c = rl_canvas.Canvas(pdf_path, pagesize=A4)
    for pil_img in pil_images:
        img_w, img_h = pil_img.size
        scale        = min(usable_w / img_w, usable_h / img_h)
        draw_w       = img_w * scale
        draw_h       = img_h * scale
        x = MARGIN + (usable_w - draw_w) / 2
        y = MARGIN + (usable_h - draw_h) / 2
        buf = io.BytesIO()
        pil_img.save(buf, format="JPEG", quality=95, optimize=True, subsampling=0)
        buf.seek(0)
        c.drawImage(ImageReader(buf), x, y, width=draw_w, height=draw_h)
        c.showPage()
    c.save()


# ================= EXCEL BUILDER =================
def build_users_excel(rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    header_fill  = PatternFill("solid", fgColor="1E3A8A")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    headers    = ["#", "User ID", "Username", "Status", "Obuna tugashi", "Ro'yxatdan o'tgan", "Takliflar"]
    col_widths = [5,   16,        24,          14,       18,              22,                   12]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font, cell.fill, cell.alignment = header_font, header_fill, center_align
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 22
    now = datetime.now()
    premium_fill = PatternFill("solid", fgColor="DCFCE7")
    normal_fill  = PatternFill("solid", fgColor="F8FAFC")
    alt_fill     = PatternFill("solid", fgColor="EFF6FF")
    premium_font = Font(name="Arial", size=10, color="166534", bold=True)
    normal_font  = Font(name="Arial", size=10, color="1E293B")
    for i, r in enumerate(rows, start=1):
        row_num    = i + 1
        is_premium = bool(r.get("access_until") and r["access_until"] > now)
        username   = f"@{r['username']}" if r.get("username") else "—"
        status     = "✅ Premium" if is_premium else "👤 Oddiy"
        access_str = r["access_until"].strftime("%d.%m.%Y") if r.get("access_until") else "—"
        reg_str    = r["created_at"].strftime("%d.%m.%Y %H:%M") if r.get("created_at") else "—"
        ref_cnt    = r.get("ref_count", 0) or 0
        row_data   = [i, r["user_id"], username, status, access_str, reg_str, ref_cnt]
        fill = premium_fill if is_premium else (alt_fill if i % 2 == 0 else normal_fill)
        font = premium_font if is_premium else normal_font
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill      = fill
            cell.font      = font
            cell.alignment = center_align if col_idx in (1, 2, 4, 5, 7) else left_align
        ws.row_dimensions[row_num].height = 18
    total_users = len(rows)
    premium_cnt = sum(1 for r in rows if r.get("access_until") and r["access_until"] > now)
    summary_row = total_users + 3
    s_fill = PatternFill("solid", fgColor="FEF9C3")
    s_font = Font(bold=True, name="Arial", size=10, color="713F12")
    for col_idx, value in [(1,"Jami:"),(2,total_users),(3,"Premium:"),(4,premium_cnt),(5,f"Sana: {now.strftime('%d.%m.%Y %H:%M')}")]:
        cell = ws.cell(row=summary_row, column=col_idx, value=value)
        cell.font, cell.fill, cell.alignment = s_font, s_fill, center_align
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
