import os
import re
import uuid
import json
import asyncio
import io
from datetime import datetime, timedelta

import numpy as np
import pymysql
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from aiogram import Bot, Dispatcher, F, types
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
    FSInputFile, BufferedInputFile
)
from dotenv import load_dotenv
from PIL import Image, ImageFilter, ImageOps, ImageEnhance
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.utils import ImageReader

# OpenCV for document edge detection
try:
    import cv2
    HAS_CV2 = True
except ImportError:
    HAS_CV2 = False

# scipy for morphological ops (optional boost)
try:
    from scipy import ndimage as ndi
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

# ================= CONFIG =================
load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")

OWNER_ID = 6875167708
ADMIN_IDS = {OWNER_ID}

PAYNET_LINK = "https://app.paynet.uz/?m=49156&i=1abfad1a-2da7-4d8d-8509-39f59b32d538"

TARIFFS = {
    30:  "1 oy – 30 000 so'm",
    90:  "3 oy – 80 000 so'm",
    180: "6 oy – 150 000 so'm"
}

ORDER_LIMIT_PER_MONTH = 10
DEADLINE_OPTIONS = ["1 kun", "2 kun", "3 kun", "5 kun", "7 kun", "10 kun"]
ORDER_TYPES = {
    "referat":  "📝 Referat",
    "mustaqil": "📘 Mustaqil ish"
}
FILETYPE_LABELS = [
    "📄 PDF",
    "📝 Word (.docx)",
    "📊 Excel (.xlsx)",
    "📊 PowerPoint (.pptx)"
]

ALL_PERMS = ["payments", "orders", "broadcast", "revoke", "users", "listusers"]

PERM_LABELS = {
    "payments":  "💳 To'lovlar",
    "orders":    "📋 Buyurtmalar",
    "broadcast": "📢 Broadcast",
    "revoke":    "🚫 Premium bekor",
    "users":     "👥 Userlar",
    "listusers": "📊 Excel ro'yxat",
}

BOT_INFO_TEXT = """ℹ️ <b>BOT HAQIDA MA'LUMOT</b>

Ushbu bot quyidagi imkoniyatlarni taqdim etadi:

━━━━━━━━━━━━━━━━━━━━━

📄 <b>PDF YARATISH</b> (bepul)
  • <b>Skan PDF</b> — Rasmlaringizni CamScanner uslubida qayta ishlaydi: oq fon, qora yozuv, professional ko'rinish
  • <b>Oddiy PDF</b> — Rasmlarni shunchaki PDF ga birlashtiradi, o'zgartirmasdan

━━━━━━━━━━━━━━━━━━━━━

💳 <b>PREMIUM OBUNA</b>
  • 1 oy — 30 000 so'm
  • 3 oy — 80 000 so'm
  • 6 oy — 150 000 so'm

━━━━━━━━━━━━━━━━━━━━━

📝 <b>REFERAT YOZDIRISH</b> (premium)
  Kerakli fandan, mavzuda, belgilangan sahifada professional referat tayyorlanadi

📘 <b>MUSTAQIL ISH YOZDIRISH</b> (premium)
  Mustaqil ish topshiriqlari tayyor holda yetkazib beriladi

━━━━━━━━━━━━━━━━━━━━━

📋 <b>BUYURTMALARIM</b>
  Barcha buyurtmalaringiz holati va tarixi

━━━━━━━━━━━━━━━━━━━━━

⏰ Buyurtma muddati: 1–10 kun
📁 Fayl formatlari: PDF, Word, Excel, PowerPoint
📊 Oylik limit: 10 ta buyurtma

━━━━━━━━━━━━━━━━━━━━━

Savollar uchun admin bilan bog'laning."""

# ================= FSM STATES =================
class PDFStates(StatesGroup):
    choosing_type     = State()
    collecting_images = State()
    waiting_pdf_name  = State()

class PaymentStates(StatesGroup):
    waiting_check = State()

class OrderStates(StatesGroup):
    entering_subject  = State()
    entering_topic    = State()
    choosing_pages    = State()
    choosing_filetype = State()
    choosing_deadline = State()
    confirming        = State()

class DeliverStates(StatesGroup):
    waiting_file = State()
    waiting_text = State()

class BroadcastStates(StatesGroup):
    waiting_message = State()
    confirming      = State()

class RevokeStates(StatesGroup):
    waiting_user_id = State()

class SubAdminStates(StatesGroup):
    waiting_user_id = State()
    choosing_perms  = State()

class SupportStates(StatesGroup):
    waiting_message = State()

class SupportReplyStates(StatesGroup):
    waiting_reply = State()

# ================= DB =================
def get_db():
    return pymysql.connect(
        host=os.getenv("DB_HOST"),
        port=int(os.getenv("DB_PORT", 3306)),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        database=os.getenv("DB_NAME"),
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=True,
        connect_timeout=10
    )

def ensure_admins_table():
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("""
                    CREATE TABLE IF NOT EXISTS admins (
                        user_id    BIGINT PRIMARY KEY,
                        username   VARCHAR(255) DEFAULT NULL,
                        perms      TEXT NOT NULL DEFAULT '[]',
                        added_at   DATETIME DEFAULT CURRENT_TIMESTAMP
                    )
                """)
        finally:
            db.close()
    except Exception as e:
        print(f"[ensure_admins_table ERROR] {e}")

# ================= PERMISSION HELPERS =================
def get_sub_admin(uid: int) -> dict | None:
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT * FROM admins WHERE user_id=%s", (uid,))
                return c.fetchone()
        finally:
            db.close()
    except Exception as e:
        print(f"[get_sub_admin ERROR] {e}")
        return None

def is_any_admin(uid: int) -> bool:
    if uid == OWNER_ID:
        return True
    return get_sub_admin(uid) is not None

def has_perm(uid: int, perm: str) -> bool:
    if uid == OWNER_ID:
        return True
    row = get_sub_admin(uid)
    if not row:
        return False
    try:
        return perm in json.loads(row["perms"])
    except Exception:
        return False

def get_admin_perms(uid: int) -> list:
    if uid == OWNER_ID:
        return list(ALL_PERMS)
    row = get_sub_admin(uid)
    if not row:
        return []
    try:
        return json.loads(row["perms"])
    except Exception:
        return []

def all_sub_admins() -> list:
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT * FROM admins ORDER BY added_at DESC")
                return c.fetchall()
        finally:
            db.close()
    except Exception as e:
        print(f"[all_sub_admins ERROR] {e}")
        return []

# ================= BOT =================
bot = Bot(token=BOT_TOKEN)
dp  = Dispatcher(storage=MemoryStorage())

# ================= MENUS =================
menu_basic = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📄 PDF yaratish")],
        [KeyboardButton(text="💳 To'lov qilish"), KeyboardButton(text="✅ To'lov qildim")],
        [KeyboardButton(text="ℹ️ Ma'lumot"),       KeyboardButton(text="🆘 Yordam")]
    ],
    resize_keyboard=True
)

menu_premium = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📄 PDF yaratish")],
        [KeyboardButton(text="📝 Referat yozdirish")],
        [KeyboardButton(text="📘 Mustaqil ish yozdirish")],
        [KeyboardButton(text="📋 Buyurtmalarim")],
        [KeyboardButton(text="ℹ️ Ma'lumot"),       KeyboardButton(text="🆘 Yordam")]
    ],
    resize_keyboard=True
)

def menu_pdf_collecting(count: int, pdf_mode: str) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📥 PDF yaratish")],
            [KeyboardButton(text="🗑 Tozalash")],
            [KeyboardButton(text="❌ Bekor qilish")]
        ],
        resize_keyboard=True
    )

# ================= ADMIN PANEL KEYBOARDS =================
def admin_panel_kb(uid: int) -> InlineKeyboardMarkup:
    rows = []
    row = []
    if has_perm(uid, "users"):
        row.append(InlineKeyboardButton(text="👥 Userlar", callback_data="adm:users"))
    if has_perm(uid, "payments"):
        row.append(InlineKeyboardButton(text="💳 To'lovlar", callback_data="adm:payments"))
    if row:
        rows.append(row)

    row2 = []
    if has_perm(uid, "orders"):
        row2.append(InlineKeyboardButton(text="⏰ Tugayotganlar", callback_data="adm:expiring"))
        row2.append(InlineKeyboardButton(text="📋 Buyurtmalar",   callback_data="adm:orders"))
    if row2:
        rows.append(row2)

    row3 = []
    if has_perm(uid, "broadcast"):
        row3.append(InlineKeyboardButton(text="📢 Broadcast", callback_data="adm:broadcast"))
    if has_perm(uid, "revoke"):
        row3.append(InlineKeyboardButton(text="🚫 Premium bekor", callback_data="adm:revoke"))
    if row3:
        rows.append(row3)

    if has_perm(uid, "payments"):
        rows.append([InlineKeyboardButton(text="⏳ Kutilayotgan to'lovlar", callback_data="adm:pending_payments")])

    if has_perm(uid, "listusers"):
        rows.append([InlineKeyboardButton(text="📊 Userlar ro'yxati (Excel)", callback_data="adm:listusers")])

    if uid == OWNER_ID:
        rows.append([InlineKeyboardButton(text="🛡 Sub-adminlar", callback_data="adm:subadmins")])

    rows.append([InlineKeyboardButton(text="🔄 Yangilash", callback_data="adm:refresh")])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def perm_toggle_kb(target_uid: int, current_perms: list) -> InlineKeyboardMarkup:
    rows = []
    for perm in ALL_PERMS:
        check = "✅" if perm in current_perms else "☐"
        rows.append([InlineKeyboardButton(
            text=f"{check} {PERM_LABELS[perm]}",
            callback_data=f"permToggle:{target_uid}:{perm}"
        )])
    rows.append([
        InlineKeyboardButton(text="💾 Saqlash",       callback_data=f"permSave:{target_uid}"),
        InlineKeyboardButton(text="❌ Bekor qilish",  callback_data="permCancel"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def subadmins_list_kb(admins: list) -> InlineKeyboardMarkup:
    rows = []
    for a in admins:
        uname = f"@{a['username']}" if a.get("username") else f"ID:{a['user_id']}"
        try:
            cnt = len(json.loads(a["perms"]))
        except Exception:
            cnt = 0
        rows.append([InlineKeyboardButton(
            text=f"🛡 {uname} ({cnt} ruxsat)",
            callback_data=f"subadmView:{a['user_id']}"
        )])
    rows.append([InlineKeyboardButton(text="➕ Yangi sub-admin qo'shish", callback_data="subadmAdd")])
    rows.append([InlineKeyboardButton(text="🔙 Orqaga", callback_data="adm:refresh")])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def subadmin_detail_kb(target_uid: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✏️ Ruxsatlarni o'zgartirish", callback_data=f"subadmEdit:{target_uid}"),
            InlineKeyboardButton(text="🗑 O'chirish",                 callback_data=f"subadmDel:{target_uid}"),
        ],
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="adm:subadmins")],
    ])

# ================= OTHER KEYBOARDS =================
def kb_pdf_type() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="🔬 Skan PDF", callback_data="pdftype:scan"),
            InlineKeyboardButton(text="📄 Oddiy PDF", callback_data="pdftype:simple"),
        ]
    ])

def kb_pages():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="5"),  KeyboardButton(text="10"), KeyboardButton(text="15")],
            [KeyboardButton(text="20"), KeyboardButton(text="25")],
            [KeyboardButton(text="❌ Bekor qilish")]
        ], resize_keyboard=True
    )

def kb_filetype():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📄 PDF"),           KeyboardButton(text="📝 Word (.docx)")],
            [KeyboardButton(text="📊 Excel (.xlsx)"), KeyboardButton(text="📊 PowerPoint (.pptx)")],
            [KeyboardButton(text="❌ Bekor qilish")]
        ], resize_keyboard=True
    )

def kb_deadline():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=d) for d in DEADLINE_OPTIONS[:3]],
            [KeyboardButton(text=d) for d in DEADLINE_OPTIONS[3:]],
            [KeyboardButton(text="❌ Bekor qilish")]
        ], resize_keyboard=True
    )

def kb_confirm():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="✅ Tasdiqlash")], [KeyboardButton(text="❌ Bekor qilish")]],
        resize_keyboard=True
    )

def kb_cancel():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="❌ Bekor qilish")]],
        resize_keyboard=True
    )

def kb_broadcast_confirm():
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Yuborish",     callback_data="bc:send"),
        InlineKeyboardButton(text="❌ Bekor qilish", callback_data="bc:cancel"),
    ]])

# ================= HELPERS =================
def has_access(uid: int) -> bool:
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT access_until FROM users WHERE user_id=%s", (uid,))
                row = c.fetchone()
            return bool(row and row["access_until"] and row["access_until"] > datetime.now())
        finally:
            db.close()
    except Exception as e:
        print(f"[has_access ERROR] {e}")
        return False

def get_menu(uid: int):
    return menu_premium if has_access(uid) else menu_basic

def sanitize_filename(name: str) -> str:
    name = re.sub(r'[^\w\s\-]', '', name, flags=re.UNICODE).strip()
    return name or "document"

def cleanup_files(*paths):
    for p in paths:
        if p and os.path.exists(p):
            try:
                os.remove(p)
            except OSError:
                pass

def format_deadline(deadline_val) -> str:
    if not deadline_val:
        return "—"
    if hasattr(deadline_val, "strftime"):
        return deadline_val.strftime("%d.%m.%Y")
    return str(deadline_val)

def get_order_monthly_count(uid: int) -> int:
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                c.execute("SELECT COUNT(*) AS c FROM orders WHERE user_id=%s AND created_at >= %s", (uid, month_start))
                return c.fetchone()["c"]
        finally:
            db.close()
    except Exception as e:
        print(f"[get_order_monthly_count ERROR] {e}")
        return 0

def get_active_order(uid: int):
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute(
                    "SELECT id, type, subject, topic, status FROM orders "
                    "WHERE user_id=%s AND status IN ('pending','in_progress') ORDER BY id DESC LIMIT 1",
                    (uid,)
                )
                return c.fetchone()
        finally:
            db.close()
    except Exception as e:
        print(f"[get_active_order ERROR] {e}")
        return None

def deadline_str_to_date(deadline_str: str) -> str:
    try:
        days = int(deadline_str.split()[0])
        return (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
    except (ValueError, IndexError):
        return (datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d")

def get_all_user_ids() -> list:
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT user_id FROM users")
                return [r["user_id"] for r in c.fetchall()]
        finally:
            db.close()
    except Exception as e:
        print(f"[get_all_user_ids ERROR] {e}")
        return []

def get_admin_stats() -> dict:
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT COUNT(*) AS c FROM users")
                total = c.fetchone()["c"]
                c.execute("SELECT COUNT(*) AS c FROM users WHERE access_until > %s", (datetime.now(),))
                premium = c.fetchone()["c"]
                c.execute("SELECT COUNT(*) AS c FROM payments WHERE status='pending'")
                pending_pay = c.fetchone()["c"]
                c.execute("SELECT COUNT(*) AS c FROM orders WHERE status IN ('pending','in_progress')")
                active_orders = c.fetchone()["c"]
            return {"total": total, "premium": premium, "pending_pay": pending_pay, "active_orders": active_orders}
        finally:
            db.close()
    except Exception as e:
        print(f"[get_admin_stats ERROR] {e}")
        return {"total": 0, "premium": 0, "pending_pay": 0, "active_orders": 0}

def build_admin_text(stats: dict, uid: int) -> str:
    role = "👑 OWNER" if uid == OWNER_ID else "🛡 SUB-ADMIN"
    sub_cnt = len(all_sub_admins())
    base = (
        f"📊 <b>ADMIN PANEL</b>  [{role}]\n\n"
        f"👤 Jami userlar:        <b>{stats['total']}</b>\n"
        f"✅ Premium:             <b>{stats['premium']}</b>\n"
        f"⏳ Kutilayotgan to'lov: <b>{stats['pending_pay']}</b>\n"
        f"📋 Aktiv buyurtmalar:   <b>{stats['active_orders']}</b>\n"
    )
    if uid == OWNER_ID:
        base += f"🛡 Sub-adminlar:        <b>{sub_cnt}</b>\n"
    base += f"\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    return base

def user_submitted_check_today(uid: int) -> bool:
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                since = datetime.now() - timedelta(hours=24)
                c.execute(
                    "SELECT COUNT(*) AS c FROM payments WHERE user_id=%s AND created_at >= %s",
                    (uid, since)
                )
                return c.fetchone()["c"] > 0
        finally:
            db.close()
    except Exception as e:
        print(f"[user_submitted_check_today ERROR] {e}")
        return False


# ================================================================
# ══════════  AUTO-CROP DOCUMENT EDGES (CamScanner style)  ═══════
# ================================================================

def order_points(pts: np.ndarray) -> np.ndarray:
    """Order 4 corner points: top-left, top-right, bottom-right, bottom-left."""
    rect = np.zeros((4, 2), dtype=np.float32)
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]   # top-left
    rect[2] = pts[np.argmax(s)]   # bottom-right
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]  # top-right
    rect[3] = pts[np.argmax(diff)]  # bottom-left
    return rect


def four_point_transform(image: np.ndarray, pts: np.ndarray) -> np.ndarray:
    """Perspective-correct warp to a top-down rectangle."""
    rect = order_points(pts)
    (tl, tr, br, bl) = rect

    widthA  = np.linalg.norm(br - bl)
    widthB  = np.linalg.norm(tr - tl)
    maxW    = max(int(widthA), int(widthB))

    heightA = np.linalg.norm(tr - br)
    heightB = np.linalg.norm(tl - bl)
    maxH    = max(int(heightA), int(heightB))

    dst = np.array([
        [0, 0],
        [maxW - 1, 0],
        [maxW - 1, maxH - 1],
        [0, maxH - 1]
    ], dtype=np.float32)

    M       = cv2.getPerspectiveTransform(rect, dst)
    warped  = cv2.warpPerspective(image, M, (maxW, maxH))
    return warped


def auto_crop_document(pil_img: Image.Image) -> Image.Image:
    """
    Detect the document boundary and perspective-correct warp it.
    Falls back to the original image if no clear quadrilateral is found.
    Requires OpenCV.
    """
    if not HAS_CV2:
        return pil_img

    orig = np.array(pil_img.convert("RGB"))
    h, w = orig.shape[:2]

    # ── Downscale for fast contour detection ──────────────────────
    scale     = 800 / max(h, w)
    small     = cv2.resize(orig, (int(w * scale), int(h * scale)))
    gray      = cv2.cvtColor(small, cv2.COLOR_RGB2GRAY)

    # ── Bilateral filter keeps edges sharp while smoothing texture ─
    blurred   = cv2.bilateralFilter(gray, 9, 75, 75)

    # ── Canny edge detection with auto thresholds ─────────────────
    med       = float(np.median(blurred))
    lo, hi    = int(max(0, 0.66 * med)), int(min(255, 1.33 * med))
    edges     = cv2.Canny(blurred, lo, hi)
    edges     = cv2.dilate(edges, np.ones((3, 3), np.uint8), iterations=1)

    # ── Find the largest 4-point contour ─────────────────────────
    cnts, _   = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    cnts      = sorted(cnts, key=cv2.contourArea, reverse=True)[:10]

    screen_cnt = None
    for c in cnts:
        peri   = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)
        if len(approx) == 4:
            area = cv2.contourArea(approx)
            # Must cover at least 15% of the small image
            if area > 0.15 * small.shape[0] * small.shape[1]:
                screen_cnt = approx
                break

    if screen_cnt is None:
        return pil_img  # No document found – return as-is

    # ── Scale corners back to original resolution ─────────────────
    pts      = screen_cnt.reshape(4, 2).astype(np.float32)
    pts     /= scale

    warped   = four_point_transform(orig, pts)
    return Image.fromarray(warped)


# ================================================================
# ══════════  PROFESSIONAL SCAN PROCESSING PIPELINE  ═════════════
# ================================================================

def process_scan_image(img: Image.Image) -> Image.Image:
    """
    Full CamScanner-quality pipeline:
      1. Auto-crop / perspective-correct
      2. Upscale for processing quality
      3. Denoise
      4. Adaptive threshold (eliminates ALL shadows)
      5. Morphological cleanup
      6. Sharpen text edges
      7. Final white-paper / black-ink binarization
    """
    # ── Step 1: Auto-crop document edges ──────────────────────────
    img = auto_crop_document(img)

    # ── Step 2: Convert to grayscale & upscale for quality ─────────
    img = img.convert("L")
    orig_w, orig_h = img.size

    # Upscale to at least 2400px on the long side for crisp output
    long_side = max(orig_w, orig_h)
    if long_side < 2400:
        factor = 2400 / long_side
        img = img.resize(
            (int(orig_w * factor), int(orig_h * factor)),
            Image.LANCZOS
        )

    arr = np.array(img, dtype=np.uint8)

    if HAS_CV2:
        # ── Step 3a: Denoise with OpenCV fastNlMeans ─────────────
        arr = cv2.fastNlMeansDenoising(arr, h=10, templateWindowSize=7, searchWindowSize=21)

        # ── Step 4a: CLAHE contrast enhancement ───────────────────
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        arr   = clahe.apply(arr)

        # ── Step 5a: Adaptive threshold — the key shadow killer ───
        # Uses a large block size so illumination gradients are fully compensated
        binary = cv2.adaptiveThreshold(
            arr,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            blockSize=51,   # large block catches wide shadow gradients
            C=18            # aggressive offset for crisp black ink
        )

        # ── Step 6a: Morphological cleanup (remove noise dots) ────
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
        binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN,  kernel)

        # ── Step 7a: Unsharp mask for crisp text edges ────────────
        blurred_sharp = cv2.GaussianBlur(binary, (0, 0), 1.2)
        binary        = cv2.addWeighted(binary, 1.8, blurred_sharp, -0.8, 0)
        _, binary     = cv2.threshold(binary, 127, 255, cv2.THRESH_BINARY)

        result_arr = binary

    else:
        # ── Fallback (no OpenCV): pure NumPy/PIL pipeline ─────────

        # Step 3b: Gaussian blur for noise reduction
        blurred = Image.fromarray(arr).filter(ImageFilter.GaussianBlur(radius=1))
        arr     = np.array(blurred, dtype=np.float32)

        # Step 4b: Enhance contrast
        p2, p98 = np.percentile(arr, 2), np.percentile(arr, 98)
        if p98 > p2:
            arr = np.clip((arr - p2) / (p98 - p2) * 255, 0, 255)

        # Step 5b: Local adaptive threshold via large-window background estimate
        arr_u8  = arr.astype(np.uint8)
        bg_img  = Image.fromarray(arr_u8).filter(ImageFilter.GaussianBlur(radius=40))
        bg      = np.array(bg_img, dtype=np.float32)
        # Subtract illumination — any pixel darker than background by threshold → ink
        diff    = bg.astype(np.float32) - arr.astype(np.float32)
        binary  = np.where(diff > 20, 0, 255).astype(np.uint8)

        # Step 6b: Cleanup with a second threshold pass
        binary  = np.where(binary < 128, 0, 255).astype(np.uint8)

        # Step 7b: Sharpening via PIL
        sharp_img = Image.fromarray(binary)
        sharp_img = sharp_img.filter(
            ImageFilter.UnsharpMask(radius=1, percent=200, threshold=2)
        )
        binary    = np.where(np.array(sharp_img) < 128, 0, 255).astype(np.uint8)

        result_arr = binary

    # ── Final: pure B&W — make whites brilliant ────────────────────
    # Any gray residue ≥ 220 → pure white; ≤ 80 → pure black
    result_arr = np.where(result_arr >= 220, 255,
                 np.where(result_arr <= 80,    0, result_arr)).astype(np.uint8)

    result_img = Image.fromarray(result_arr, mode="L")

    # ── Compose on RGB white canvas ────────────────────────────────
    out = Image.new("RGB", result_img.size, (255, 255, 255))
    out.paste(result_img)
    return out


# ================================================================
# ═════════════  PDF BUILDER — Uniform A4 pages  ═════════════════
# ================================================================

def build_pdf_with_reportlab(pil_images: list, pdf_path: str):
    PAGE_W, PAGE_H = A4
    MARGIN   = 20          # tighter margin for scan PDFs — more paper coverage
    usable_w = PAGE_W - 2 * MARGIN
    usable_h = PAGE_H - 2 * MARGIN
    c = rl_canvas.Canvas(pdf_path, pagesize=A4)
    for pil_img in pil_images:
        img_w, img_h = pil_img.size
        scale        = min(usable_w / img_w, usable_h / img_h)
        draw_w       = img_w * scale
        draw_h       = img_h * scale
        x = MARGIN + (usable_w - draw_w) / 2
        y = MARGIN + (usable_h - draw_h) / 2
        buf = io.BytesIO()
        # Use high-quality compression for scan images
        pil_img.save(buf, format="JPEG", quality=95, optimize=True, subsampling=0)
        buf.seek(0)
        c.drawImage(ImageReader(buf), x, y, width=draw_w, height=draw_h)
        c.showPage()
    c.save()


# ================= EXCEL BUILDER =================
def build_users_excel(rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    header_fill  = PatternFill("solid", fgColor="1E3A8A")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    headers    = ["#", "User ID", "Username", "Status", "Obuna tugashi", "Ro'yxatdan o'tgan"]
    col_widths = [5,   16,        24,          14,       18,              22]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font, cell.fill, cell.alignment = header_font, header_fill, center_align
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 22
    now = datetime.now()
    premium_fill = PatternFill("solid", fgColor="DCFCE7")
    normal_fill  = PatternFill("solid", fgColor="F8FAFC")
    alt_fill     = PatternFill("solid", fgColor="EFF6FF")
    premium_font = Font(name="Arial", size=10, color="166534", bold=True)
    normal_font  = Font(name="Arial", size=10, color="1E293B")
    for i, r in enumerate(rows, start=1):
        row_num    = i + 1
        is_premium = bool(r.get("access_until") and r["access_until"] > now)
        username   = f"@{r['username']}" if r.get("username") else "—"
        status     = "✅ Premium" if is_premium else "👤 Oddiy"
        access_str = r["access_until"].strftime("%d.%m.%Y") if r.get("access_until") else "—"
        reg_str    = r["created_at"].strftime("%d.%m.%Y %H:%M") if r.get("created_at") else "—"
        row_data   = [i, r["user_id"], username, status, access_str, reg_str]
        fill = premium_fill if is_premium else (alt_fill if i % 2 == 0 else normal_fill)
        font = premium_font if is_premium else normal_font
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill      = fill
            cell.font      = font
            cell.alignment = center_align if col_idx in (1, 2, 4, 5) else left_align
        ws.row_dimensions[row_num].height = 18
    total_users = len(rows)
    premium_cnt = sum(1 for r in rows if r.get("access_until") and r["access_until"] > now)
    summary_row = total_users + 3
    s_fill = PatternFill("solid", fgColor="FEF9C3")
    s_font = Font(bold=True, name="Arial", size=10, color="713F12")
    for col_idx, value in [(1,"Jami:"),(2,total_users),(3,"Premium:"),(4,premium_cnt),(5,f"Sana: {now.strftime('%d.%m.%Y %H:%M')}")]:
        cell = ws.cell(row=summary_row, column=col_idx, value=value)
        cell.font, cell.fill, cell.alignment = s_font, s_fill, center_align
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ================================================================
# ═════════════════  INFO BUTTON  ════════════════════════════════
# ================================================================

@dp.message(F.text == "ℹ️ Ma'lumot")
async def info_handler(message: types.Message):
    await message.answer(BOT_INFO_TEXT, parse_mode="HTML")

# ================================================================
# ══════════════  SUPPORT / YORDAM  ══════════════════════════════
# ================================================================

@dp.message(F.text == "🆘 Yordam")
async def support_start(message: types.Message, state: FSMContext):
    await state.set_state(SupportStates.waiting_message)
    await message.answer(
        "✍️ Xabaringizni yozing:",
        reply_markup=kb_cancel()
    )

@dp.message(SupportStates.waiting_message, F.text)
async def support_receive(message: types.Message, state: FSMContext):
    uid      = message.from_user.id
    username = message.from_user.username or "—"
    text     = message.text.strip()
    await state.clear()

    admin_txt = (
        f"💬 <b>YORDAM SO'ROVI</b>\n"
        f"👤 @{username} (ID: <code>{uid}</code>)\n"
        f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
        f"{'─'*25}\n\n"
        f"{text}"
    )
    reply_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="↩️ Javob berish", callback_data=f"supReply:{uid}")
    ]])
    try:
        await bot.send_message(OWNER_ID, admin_txt, reply_markup=reply_kb, parse_mode="HTML")
    except Exception as e:
        print(f"[SUPPORT FORWARD ERROR] {e}")

    await message.answer(
        "✅ Xabaringiz adminga yuborildi!",
        reply_markup=get_menu(uid)
    )

@dp.callback_query(F.data.startswith("supReply:"))
async def support_reply_start(callback: types.CallbackQuery, state: FSMContext):
    if callback.from_user.id != OWNER_ID:
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    target_uid = int(callback.data.split(":")[1])
    await state.set_state(SupportReplyStates.waiting_reply)
    await state.update_data(support_target_uid=target_uid)
    await callback.answer()
    await callback.message.reply(
        f"✏️ <code>{target_uid}</code> foydalanuvchiga javobingizni yozing:",
        reply_markup=kb_cancel(),
        parse_mode="HTML"
    )

@dp.message(SupportReplyStates.waiting_reply, F.text)
async def support_reply_send(message: types.Message, state: FSMContext):
    data       = await state.get_data()
    target_uid = data.get("support_target_uid")
    await state.clear()
    try:
        await bot.send_message(
            target_uid,
            f"<b>Admin:</b>\n\n{message.text}",
            parse_mode="HTML"
        )
        await message.answer("✅ Javob yuborildi.", reply_markup=get_menu(message.from_user.id))
    except Exception as e:
        await message.answer(f"❌ Xato: {e}")

# ================================================================
# ═══════════════  SUB-ADMIN MANAGEMENT  ═════════════════════════
# ================================================================

@dp.message(Command("admins"))
@dp.callback_query(F.data == "adm:subadmins")
async def subadmins_panel(event: types.Message | types.CallbackQuery):
    uid = event.from_user.id
    if uid != OWNER_ID:
        if isinstance(event, types.CallbackQuery):
            return await event.answer("❌ Faqat owner uchun", show_alert=True)
        return

    admins = all_sub_admins()
    text = (
        f"🛡 <b>SUB-ADMINLAR</b>\n\nJami: <b>{len(admins)}</b> ta sub-admin\n\n"
        + ("\n".join(
            f"• {'@'+a['username'] if a.get('username') else 'ID:'+str(a['user_id'])}  "
            f"({len(json.loads(a['perms']))} ruxsat)"
            for a in admins
        ) if admins else "Hozircha sub-admin yo'q.")
    )
    kb = subadmins_list_kb(admins)
    if isinstance(event, types.CallbackQuery):
        await event.answer()
        try:
            await event.message.edit_text(text, reply_markup=kb, parse_mode="HTML")
        except Exception:
            await event.message.answer(text, reply_markup=kb, parse_mode="HTML")
    else:
        await event.answer(text, reply_markup=kb, parse_mode="HTML")


@dp.callback_query(F.data.startswith("subadmView:"))
async def subadm_view(callback: types.CallbackQuery):
    if callback.from_user.id != OWNER_ID:
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    target_uid = int(callback.data.split(":")[1])
    row = get_sub_admin(target_uid)
    if not row:
        return await callback.answer("❌ Topilmadi", show_alert=True)
    try:
        perms = json.loads(row["perms"])
    except Exception:
        perms = []
    uname    = f"@{row['username']}" if row.get("username") else f"ID:{target_uid}"
    added_at = row["added_at"].strftime('%d.%m.%Y %H:%M') if row.get("added_at") else "—"
    perm_txt = "\n".join(f"  {'✅' if p in perms else '☐'} {PERM_LABELS.get(p, p)}" for p in ALL_PERMS)
    text = (
        f"🛡 <b>SUB-ADMIN: {uname}</b>\n"
        f"🆔 ID: <code>{target_uid}</code>\n"
        f"📅 Qo'shilgan: {added_at}\n\n"
        f"<b>Ruxsatlar:</b>\n{perm_txt}"
    )
    await callback.answer()
    try:
        await callback.message.edit_text(text, reply_markup=subadmin_detail_kb(target_uid), parse_mode="HTML")
    except Exception:
        await callback.message.answer(text, reply_markup=subadmin_detail_kb(target_uid), parse_mode="HTML")


@dp.callback_query(F.data == "subadmAdd")
async def subadm_add_start(callback: types.CallbackQuery, state: FSMContext):
    if callback.from_user.id != OWNER_ID:
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    await callback.answer()
    await state.set_state(SubAdminStates.waiting_user_id)
    await callback.message.answer(
        "🛡 <b>YANGI SUB-ADMIN</b>\n\nAdmin qilmoqchi bo'lgan foydalanuvchining <b>User ID</b>sini yuboring:\n<i>Masalan: 123456789</i>",
        reply_markup=kb_cancel(), parse_mode="HTML"
    )


@dp.message(SubAdminStates.waiting_user_id, F.text)
async def subadm_got_uid(message: types.Message, state: FSMContext):
    if message.from_user.id != OWNER_ID:
        return
    text = message.text.strip()
    if not text.lstrip("-").isdigit():
        return await message.answer("❌ Noto'g'ri format. Faqat raqam kiriting:")
    target_uid = int(text)
    if target_uid == OWNER_ID:
        return await message.answer("❌ O'zingizni sub-admin qila olmaysiz.")

    existing = get_sub_admin(target_uid)
    current_perms = []
    if existing:
        try:
            current_perms = json.loads(existing["perms"])
        except Exception:
            current_perms = []

    username = None
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT username FROM users WHERE user_id=%s", (target_uid,))
                row = c.fetchone()
                if row:
                    username = row["username"]
        finally:
            db.close()
    except Exception:
        pass

    await state.set_state(SubAdminStates.choosing_perms)
    await state.update_data(target_uid=target_uid, current_perms=current_perms, target_username=username)
    uname_str = f"@{username}" if username else f"ID:{target_uid}"
    action    = "tahrirlash" if existing else "qo'shish"
    await message.answer(
        f"🛡 <b>Sub-admin {action}: {uname_str}</b>\n\nQuyidagi ruxsatlarni tanlang.\nTayyor bo'lgach <b>💾 Saqlash</b>ni bosing.",
        reply_markup=perm_toggle_kb(target_uid, current_perms), parse_mode="HTML"
    )


@dp.callback_query(F.data.startswith("permToggle:"))
async def perm_toggle(callback: types.CallbackQuery, state: FSMContext):
    if callback.from_user.id != OWNER_ID:
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    _, target_uid_str, perm = callback.data.split(":")
    target_uid = int(target_uid_str)
    data = await state.get_data()
    current_perms = list(data.get("current_perms", []))
    if perm in current_perms:
        current_perms.remove(perm)
    else:
        current_perms.append(perm)
    await state.update_data(current_perms=current_perms)
    try:
        await callback.message.edit_reply_markup(reply_markup=perm_toggle_kb(target_uid, current_perms))
    except Exception:
        pass
    await callback.answer(f"{'✅ Yoqildi' if perm in current_perms else '☐ O`chirildi'}: {PERM_LABELS.get(perm, perm)}")


@dp.callback_query(F.data.startswith("permSave:"))
async def perm_save(callback: types.CallbackQuery, state: FSMContext):
    if callback.from_user.id != OWNER_ID:
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    target_uid = int(callback.data.split(":")[1])
    data = await state.get_data()
    current_perms   = data.get("current_perms", [])
    target_username = data.get("target_username")
    perms_json = json.dumps(current_perms)
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute(
                    """INSERT INTO admins (user_id, username, perms, added_at)
                       VALUES (%s, %s, %s, %s)
                       ON DUPLICATE KEY UPDATE username=%s, perms=%s""",
                    (target_uid, target_username, perms_json, datetime.now(), target_username, perms_json)
                )
        finally:
            db.close()
    except Exception as e:
        print(f"[PERM SAVE ERROR] {e}")
        await callback.answer("❌ DB xato", show_alert=True)
        await state.clear()
        return

    await state.clear()
    uname_str = f"@{target_username}" if target_username else f"ID:{target_uid}"
    perm_txt  = ", ".join(PERM_LABELS.get(p, p) for p in current_perms) or "Hech biri"
    await callback.answer("✅ Saqlandi")
    try:
        await callback.message.edit_text(
            f"✅ <b>Sub-admin saqlandi!</b>\n\n👤 {uname_str}\n🔑 Ruxsatlar: {perm_txt}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="🔙 Sub-adminlar", callback_data="adm:subadmins")
            ]]),
            parse_mode="HTML"
        )
    except Exception:
        await callback.message.answer(f"✅ Sub-admin saqlandi: {uname_str}\nRuxsatlar: {perm_txt}", parse_mode="HTML")

    try:
        perm_list = "\n".join(f"  • {PERM_LABELS.get(p,p)}" for p in current_perms) or "  • Hech biri"
        await bot.send_message(
            target_uid,
            f"🛡 <b>Siz sub-admin sifatida tayinlandingiz!</b>\n\nSizga berilgan ruxsatlar:\n{perm_list}\n\nAdmin panelni ochish uchun /admin buyrug'ini yuboring.",
            parse_mode="HTML"
        )
    except Exception as e:
        print(f"[NOTIFY SUBADMIN ERROR] {e}")


@dp.callback_query(F.data == "permCancel")
async def perm_cancel(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.answer("❌ Bekor qilindi")
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass
    await callback.message.answer("❌ Bekor qilindi.")


@dp.callback_query(F.data.startswith("subadmEdit:"))
async def subadm_edit(callback: types.CallbackQuery, state: FSMContext):
    if callback.from_user.id != OWNER_ID:
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    target_uid = int(callback.data.split(":")[1])
    row = get_sub_admin(target_uid)
    if not row:
        return await callback.answer("❌ Topilmadi", show_alert=True)
    try:
        current_perms = json.loads(row["perms"])
    except Exception:
        current_perms = []
    await state.set_state(SubAdminStates.choosing_perms)
    await state.update_data(target_uid=target_uid, current_perms=current_perms, target_username=row.get("username"))
    uname_str = f"@{row['username']}" if row.get("username") else f"ID:{target_uid}"
    await callback.answer()
    try:
        await callback.message.edit_text(
            f"✏️ <b>Ruxsatlarni o'zgartirish: {uname_str}</b>\n\nRuxsatlarni yoqing yoki o'chiring, so'ng 💾 Saqlash.",
            reply_markup=perm_toggle_kb(target_uid, current_perms), parse_mode="HTML"
        )
    except Exception:
        await callback.message.answer(
            f"✏️ Ruxsatlarni o'zgartirish: {uname_str}",
            reply_markup=perm_toggle_kb(target_uid, current_perms), parse_mode="HTML"
        )


@dp.callback_query(F.data.startswith("subadmDel:"))
async def subadm_delete(callback: types.CallbackQuery):
    if callback.from_user.id != OWNER_ID:
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    target_uid = int(callback.data.split(":")[1])
    row = get_sub_admin(target_uid)
    if not row:
        return await callback.answer("❌ Topilmadi", show_alert=True)
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("DELETE FROM admins WHERE user_id=%s", (target_uid,))
        finally:
            db.close()
    except Exception as e:
        print(f"[SUBADM DELETE ERROR] {e}")
        return await callback.answer("❌ DB xato", show_alert=True)

    uname_str = f"@{row['username']}" if row.get("username") else f"ID:{target_uid}"
    await callback.answer(f"✅ {uname_str} o'chirildi")
    try:
        await bot.send_message(target_uid, "⚠️ Sizning admin huquqlaringiz bekor qilindi.")
    except Exception:
        pass
    admins = all_sub_admins()
    text = (
        f"🛡 <b>SUB-ADMINLAR</b>\n\nJami: <b>{len(admins)}</b> ta sub-admin\n\n"
        + ("\n".join(
            f"• {'@'+a['username'] if a.get('username') else 'ID:'+str(a['user_id'])}  "
            f"({len(json.loads(a['perms']))} ruxsat)"
            for a in admins
        ) if admins else "Hozircha sub-admin yo'q.")
    )
    try:
        await callback.message.edit_text(text, reply_markup=subadmins_list_kb(admins), parse_mode="HTML")
    except Exception:
        await callback.message.answer(text, reply_markup=subadmins_list_kb(admins), parse_mode="HTML")

# ================================================================
# ═════════════════  LISTUSERS  ══════════════════════════════════
# ================================================================

@dp.message(Command("listusers"))
@dp.callback_query(F.data == "adm:listusers")
async def listusers_cmd(event: types.Message | types.CallbackQuery):
    uid = event.from_user.id
    if not has_perm(uid, "listusers"):
        if isinstance(event, types.CallbackQuery):
            return await event.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    if isinstance(event, types.CallbackQuery):
        await event.answer()
        target = event.message
    else:
        target = event

    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT user_id, username, access_until, created_at FROM users ORDER BY created_at DESC")
                rows = c.fetchall()
        finally:
            db.close()
    except Exception as e:
        print(f"[LISTUSERS DB ERROR] {e}")
        return await target.answer("❌ DB xato.")

    if not rows:
        return await target.answer("❌ Hech qanday foydalanuvchi yo'q.")

    now         = datetime.now()
    total       = len(rows)
    premium_cnt = sum(1 for r in rows if r.get("access_until") and r["access_until"] > now)

    await target.answer(
        f"👥 <b>FOYDALANUVCHILAR RO'YXATI</b>\n\n"
        f"📊 Jami: <b>{total}</b> | ✅ Premium: <b>{premium_cnt}</b> | 👤 Oddiy: <b>{total-premium_cnt}</b>\n\n"
        f"⏳ Excel tayyorlanmoqda...", parse_mode="HTML"
    )
    try:
        excel_bytes = build_users_excel(rows)
        filename    = f"users_{now.strftime('%Y%m%d_%H%M')}.xlsx"
        await target.answer_document(
            BufferedInputFile(excel_bytes, filename=filename),
            caption=(
                f"📊 <b>Foydalanuvchilar ro'yxati</b>\n"
                f"👥 {total} | ✅ {premium_cnt} | 📅 {now.strftime('%d.%m.%Y %H:%M')}"
            ),
            parse_mode="HTML"
        )
    except Exception as e:
        await target.answer(f"❌ Excel xato: {e}")

# ================================================================
# ═════════════════  START  ══════════════════════════════════════
# ================================================================

@dp.message(CommandStart())
async def start(message: types.Message, state: FSMContext):
    uid      = message.from_user.id
    username = message.from_user.username or ""
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT user_id FROM users WHERE user_id=%s", (uid,))
                existing = c.fetchone()
                c.execute(
                    "INSERT IGNORE INTO users (user_id, username, created_at) VALUES (%s,%s,%s)",
                    (uid, username, datetime.now())
                )
            if not existing:
                try:
                    await bot.send_message(
                        OWNER_ID,
                        f"🆕 <b>YANGI FOYDALANUVCHI</b>\n"
                        f"👤 @{username or '—'} (ID: <code>{uid}</code>)\n"
                        f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                        parse_mode="HTML"
                    )
                except Exception as e:
                    print(f"[NEW USER NOTIFY ERROR] {e}")
        finally:
            db.close()
    except Exception as e:
        print(f"[START DB ERROR] {e}")
    await state.clear()
    await message.answer("👋 Xush kelibsiz!", reply_markup=get_menu(uid))

# ================= CANCEL =================
@dp.message(F.text == "❌ Bekor qilish")
async def cancel_any(message: types.Message, state: FSMContext):
    data = await state.get_data()
    for p in data.get("images", []):
        cleanup_files(p)
    await state.clear()
    await message.answer("❌ Bekor qilindi", reply_markup=get_menu(message.from_user.id))

# ================================================================
# ═════════════════  PDF FLOW  ═══════════════════════════════════
# ================================================================

@dp.message(F.text == "📄 PDF yaratish")
async def pdf_start(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state == PDFStates.collecting_images.state:
        data = await state.get_data()
        images   = data.get("images", [])
        pdf_mode = data.get("pdf_mode", "simple")
        if images:
            return await message.answer(
                f"📸 {len(images)} ta rasm bor.\nYana rasm yuboring yoki '📥 PDF yaratish' ni bosing.",
                reply_markup=menu_pdf_collecting(len(images), pdf_mode)
            )
    await message.answer(
        "📄 <b>PDF YARATISH</b>\n\nQaysi turdagi PDF kerak?\n\n"
        "🔬 <b>Skan PDF</b> — CamScanner uslubida: oq fon, qora matn, barcha soya va dog'lar yo'q\n"
        "📄 <b>Oddiy PDF</b> — Rasmlar o'zgarishsiz PDF ga birlashtiriladi",
        reply_markup=kb_pdf_type(), parse_mode="HTML"
    )


@dp.callback_query(F.data.startswith("pdftype:"))
async def pdf_type_chosen(callback: types.CallbackQuery, state: FSMContext):
    pdf_mode = callback.data.split(":")[1]
    await state.set_state(PDFStates.collecting_images)
    await state.update_data(images=[], pdf_mode=pdf_mode)
    mode_label = "🔬 Skan PDF" if pdf_mode == "scan" else "📄 Oddiy PDF"
    await callback.answer()
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer(
        f"✅ Rejim: <b>{mode_label}</b>\n\n📸 Rasmlarni yuboring.\nHammasi tayyor bo'lgach '📥 PDF yaratish' tugmasini bosing.",
        reply_markup=menu_pdf_collecting(0, pdf_mode), parse_mode="HTML"
    )


@dp.message(PDFStates.collecting_images, F.photo)
async def pdf_add_image(message: types.Message, state: FSMContext):
    uid  = message.from_user.id
    file = await bot.get_file(message.photo[-1].file_id)
    path = f"/tmp/img_{uid}_{uuid.uuid4().hex}.jpg"
    await bot.download_file(file.file_path, path)
    data   = await state.get_data()
    images = data.get("images", [])
    images.append(path)
    await state.update_data(images=images)
    pdf_mode = data.get("pdf_mode", "simple")
    await message.answer(
        f"✅ Rasm qo'shildi! Jami: {len(images)} ta\nYana rasm yuboring yoki '📥 PDF yaratish' ni bosing.",
        reply_markup=menu_pdf_collecting(len(images), pdf_mode)
    )


@dp.message(PDFStates.collecting_images, F.text == "🗑 Tozalash")
async def pdf_clear_images(message: types.Message, state: FSMContext):
    data   = await state.get_data()
    images = data.get("images", [])
    pdf_mode = data.get("pdf_mode", "simple")
    for p in images:
        cleanup_files(p)
    await state.update_data(images=[])
    await message.answer(
        "🗑 Barcha rasmlar o'chirildi.\nQaytadan rasm yuborishingiz mumkin.",
        reply_markup=menu_pdf_collecting(0, pdf_mode)
    )


@dp.message(PDFStates.collecting_images, F.text == "📥 PDF yaratish")
async def pdf_ask_name(message: types.Message, state: FSMContext):
    data   = await state.get_data()
    images = data.get("images", [])
    pdf_mode = data.get("pdf_mode", "simple")
    if not images:
        return await message.answer("❌ Hali rasm yuborilmagan.", reply_markup=menu_pdf_collecting(0, pdf_mode))
    await state.set_state(PDFStates.waiting_pdf_name)
    await message.answer(f"✅ {len(images)} ta rasm tayyor.\n\n📝 PDF uchun nom kiriting:", reply_markup=kb_cancel())


@dp.message(PDFStates.waiting_pdf_name, F.text)
async def pdf_create(message: types.Message, state: FSMContext):
    data     = await state.get_data()
    images   = data.get("images", [])
    pdf_mode = data.get("pdf_mode", "simple")
    pdf_path = None

    if not images:
        await state.clear()
        return await message.answer("❌ Rasm topilmadi.", reply_markup=get_menu(message.from_user.id))

    safe_name = sanitize_filename(message.text)
    pdf_path  = f"/tmp/{safe_name}_{uuid.uuid4().hex}.pdf"

    mode_label = "🔬 Skan" if pdf_mode == "scan" else "📄 Oddiy"
    await message.answer(f"⏳ {mode_label} PDF yaratilmoqda...")

    try:
        pil_images = []
        for p in images:
            img = Image.open(p).convert("RGB")
            if pdf_mode == "scan":
                img = process_scan_image(img)
            pil_images.append(img)

        build_pdf_with_reportlab(pil_images, pdf_path)

        for img in pil_images:
            img.close()

        await message.answer_document(
            FSInputFile(pdf_path, filename=f"{safe_name}.pdf"),
            caption=f"✅ {mode_label} PDF tayyor! ({len(images)} ta rasm)"
        )
    except Exception as e:
        await message.answer(f"❌ Xato: {e}")
    finally:
        cleanup_files(*images, pdf_path)
        await state.clear()

    await message.answer("✅ Bajarildi!", reply_markup=get_menu(message.from_user.id))

# ================================================================
# ═════════════════  PAYMENT  ════════════════════════════════════
# ================================================================

@dp.message(F.text == "💳 To'lov qilish")
async def payment_info(message: types.Message):
    txt = "💳 <b>Tariflar:</b>\n\n" + "".join(f"• {v}\n" for v in TARIFFS.values())
    txt += f"\n👉 To'lov:\n{PAYNET_LINK},\n yoki Karta:\n 9860350147430564"
    await message.answer(txt, parse_mode="HTML")


@dp.message(F.text == "✅ To'lov qildim")
async def wait_for_check(message: types.Message, state: FSMContext):
    uid = message.from_user.id
    if user_submitted_check_today(uid) and uid != OWNER_ID:
        return await message.answer(
            "⚠️ Siz bugungi limitdan foydalandingiz, ertaga qayta bosing."
        )
    await state.set_state(PaymentStates.waiting_check)
    await message.answer("🧾 Chekni yuboring (rasm yoki PDF)")


@dp.message(PaymentStates.waiting_check, F.photo | F.document)
async def receive_check(message: types.Message, state: FSMContext):
    uid = message.from_user.id
    await state.clear()

    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute(
                    "INSERT INTO payments (user_id, status, created_at) VALUES (%s, 'pending', %s)",
                    (uid, datetime.now())
                )
                payment_id = c.lastrowid
        finally:
            db.close()
    except Exception as e:
        print(f"[PAYMENT INSERT ERROR] {e}")
        return await message.answer("❌ Xato yuz berdi.")

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ 30 kun",  callback_data=f"pApprove:{payment_id}:{uid}:30"),
            InlineKeyboardButton(text="✅ 90 kun",  callback_data=f"pApprove:{payment_id}:{uid}:90"),
            InlineKeyboardButton(text="✅ 180 kun", callback_data=f"pApprove:{payment_id}:{uid}:180"),
        ],
        [InlineKeyboardButton(text="❌ Rad etish", callback_data=f"pReject:{payment_id}:{uid}")]
    ])
    username = message.from_user.username or "—"
    caption  = (
        f"🧾 <b>TO'LOV CHEKI</b>\n👤 @{username} (ID: <code>{uid}</code>)\n"
        f"🆔 Payment ID: {payment_id}\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )

    targets = [OWNER_ID] + [a["user_id"] for a in all_sub_admins() if "payments" in json.loads(a.get("perms", "[]"))]
    for admin_target in targets:
        try:
            if message.photo:
                await bot.send_photo(admin_target, message.photo[-1].file_id, caption=caption, reply_markup=kb, parse_mode="HTML")
            else:
                await bot.send_document(admin_target, message.document.file_id, caption=caption, reply_markup=kb, parse_mode="HTML")
        except Exception as e:
            print(f"[ADMIN SEND ERROR uid={admin_target}] {e}")

    await message.answer("⏳ Chek yuborildi. 24 soat ichida javob beriladi.")


@dp.callback_query(F.data.startswith("pApprove:"))
async def approve_payment(callback: types.CallbackQuery):
    if not has_perm(callback.from_user.id, "payments"):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    try:
        _, payment_id, uid, days = callback.data.split(":")
        payment_id, uid, days = int(payment_id), int(uid), int(days)
    except Exception:
        return await callback.answer("❌ Noto'g'ri format", show_alert=True)

    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT id, status FROM payments WHERE id=%s", (payment_id,))
                pay = c.fetchone()
            if not pay:
                return await callback.answer("❌ Payment topilmadi", show_alert=True)
            if pay["status"] != "pending":
                return await callback.answer(
                    f"⚠️ Bu to'lov allaqachon {'tasdiqlangan' if pay['status']=='approved' else 'rad etilgan'}!",
                    show_alert=True
                )
            with db.cursor() as c:
                c.execute(
                    "UPDATE users SET access_until = GREATEST(COALESCE(access_until,NOW()),NOW()) + INTERVAL %s DAY, warned=0 WHERE user_id=%s",
                    (days, uid)
                )
            with db.cursor() as c:
                c.execute("UPDATE payments SET status='approved', tariff_days=%s WHERE id=%s", (days, payment_id))
        finally:
            db.close()
    except Exception as e:
        print(f"[APPROVE PAYMENT ERROR] {e}")
        return await callback.answer("❌ DB xato", show_alert=True)

    try:
        await bot.send_message(
            uid,
            f"🎉 Premium <b>{days} kun</b> faollashtirildi!\n📅 {datetime.now().strftime('%d.%m.%Y')}",
            reply_markup=menu_premium, parse_mode="HTML"
        )
    except Exception as e:
        print(f"[SEND ERROR uid={uid}] {e}")

    admin_name = callback.from_user.username or str(callback.from_user.id)
    try:
        await callback.message.edit_caption(
            caption=callback.message.caption + f"\n\n✅ <b>Tasdiqlandi</b> (@{admin_name}, +{days} kun)",
            reply_markup=None, parse_mode="HTML"
        )
    except Exception:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
    await callback.answer(f"✅ Tasdiqlandi: +{days} kun")


@dp.callback_query(F.data.startswith("pReject:"))
async def reject_payment(callback: types.CallbackQuery):
    if not has_perm(callback.from_user.id, "payments"):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    try:
        _, payment_id, uid = callback.data.split(":")
        payment_id, uid = int(payment_id), int(uid)
    except Exception:
        return await callback.answer("❌ Noto'g'ri format", show_alert=True)

    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT id, status FROM payments WHERE id=%s", (payment_id,))
                pay = c.fetchone()
            if not pay:
                return await callback.answer("❌ Payment topilmadi", show_alert=True)
            if pay["status"] != "pending":
                return await callback.answer(
                    f"⚠️ Bu to'lov allaqachon {'tasdiqlangan' if pay['status']=='approved' else 'rad etilgan'}!",
                    show_alert=True
                )
            with db.cursor() as c:
                c.execute("UPDATE payments SET status='rejected' WHERE id=%s", (payment_id,))
        finally:
            db.close()
    except Exception as e:
        print(f"[REJECT PAYMENT ERROR] {e}")
        return await callback.answer("❌ DB xato", show_alert=True)

    try:
        await bot.send_message(uid, "❌ To'lovingiz tasdiqlanmadi.\nChekni qayta tekshirib yuboring.")
    except Exception as e:
        print(f"[SEND ERROR uid={uid}] {e}")

    admin_name = callback.from_user.username or str(callback.from_user.id)
    try:
        await callback.message.edit_caption(
            caption=callback.message.caption + f"\n\n❌ <b>Rad etildi</b> (@{admin_name})",
            reply_markup=None, parse_mode="HTML"
        )
    except Exception:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
    await callback.answer("❌ Rad etildi")

# ================================================================
# ════════  PENDING PAYMENTS LIST  ═══════════════════════════════
# ================================================================

@dp.message(Command("pendingpay"))
@dp.callback_query(F.data == "adm:pending_payments")
async def adm_pending_payments(event: types.Message | types.CallbackQuery):
    uid = event.from_user.id
    if not has_perm(uid, "payments"):
        if isinstance(event, types.CallbackQuery):
            return await event.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    if isinstance(event, types.CallbackQuery):
        await event.answer()
        target = event.message
    else:
        target = event

    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute(
                    "SELECT p.id, p.user_id, p.created_at, u.username "
                    "FROM payments p LEFT JOIN users u ON p.user_id=u.user_id "
                    "WHERE p.status='pending' ORDER BY p.id DESC"
                )
                rows = c.fetchall()
        finally:
            db.close()
    except Exception as e:
        print(f"[PENDING PAY ERROR] {e}")
        return await target.answer("❌ DB xato.")

    if not rows:
        return await target.answer("✅ Hozircha kutilayotgan to'lovlar yo'q.")

    await target.answer(
        f"⏳ <b>KUTILAYOTGAN TO'LOVLAR — {len(rows)} ta</b>",
        parse_mode="HTML"
    )
    for r in rows:
        uname    = f"@{r['username']}" if r.get("username") else "—"
        date_str = r["created_at"].strftime('%d.%m.%Y %H:%M') if r.get("created_at") else "—"
        pay_uid  = r["user_id"]
        pay_id   = r["id"]
        txt = (
            f"🧾 <b>To'lov #{pay_id}</b>\n"
            f"👤 {uname} (ID: <code>{pay_uid}</code>)\n"
            f"📅 {date_str}"
        )
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ 30 kun",  callback_data=f"pApprove:{pay_id}:{pay_uid}:30"),
                InlineKeyboardButton(text="✅ 90 kun",  callback_data=f"pApprove:{pay_id}:{pay_uid}:90"),
                InlineKeyboardButton(text="✅ 180 kun", callback_data=f"pApprove:{pay_id}:{pay_uid}:180"),
            ],
            [InlineKeyboardButton(text="❌ Rad etish", callback_data=f"pReject:{pay_id}:{pay_uid}")]
        ])
        try:
            await target.answer(txt, reply_markup=kb, parse_mode="HTML")
        except Exception as e:
            print(f"[PENDING PAY SEND ERROR] {e}")

# ================================================================
# ═════════════════  ORDER FLOW  ═════════════════════════════════
# ================================================================

async def start_order(message: types.Message, state: FSMContext, order_type: str):
    uid = message.from_user.id
    if not has_access(uid):
        return await message.answer(
            "❌ Bu funksiya faqat premium foydalanuvchilar uchun.\n💳 To'lov qilish tugmasini bosing.",
            reply_markup=get_menu(uid)
        )
    active = get_active_order(uid)
    if active:
        type_name = ORDER_TYPES.get(active["type"], active["type"])
        return await message.answer(
            f"⚠️ Sizda hali bajarilmagan buyurtma bor:\n🆔 #{active['id']} | {type_name}\n"
            f"📚 {active['subject']} — {active['topic']}\n⏳ Holati: {active['status']}\n\n"
            f"Yangi buyurtma berish uchun avvalgi buyurtma tugashini kuting.",
            reply_markup=get_menu(uid)
        )
    used = get_order_monthly_count(uid)
    if used >= ORDER_LIMIT_PER_MONTH:
        return await message.answer(
            f"⚠️ Siz bu oy {ORDER_LIMIT_PER_MONTH} ta buyurtma limitidan foydalandingiz.\nKeyingi oy yangilanadi.",
            reply_markup=get_menu(uid)
        )
    remaining  = ORDER_LIMIT_PER_MONTH - used
    type_label = ORDER_TYPES.get(order_type, order_type)
    await state.set_state(OrderStates.entering_subject)
    await state.update_data(order_type=order_type)
    await message.answer(
        f"{type_label} buyurtmasi\n📊 Bu oy: {used}/{ORDER_LIMIT_PER_MONTH} (qoldi: {remaining})\n\n"
        f"📚 Fan nomini kiriting:\n<i>Masalan: Matematika, Fizika, Tarix...</i>",
        reply_markup=kb_cancel(), parse_mode="HTML"
    )


@dp.message(F.text == "📝 Referat yozdirish")
async def order_referat(message: types.Message, state: FSMContext):
    await start_order(message, state, "referat")


@dp.message(F.text == "📘 Mustaqil ish yozdirish")
async def order_mustaqil(message: types.Message, state: FSMContext):
    await start_order(message, state, "mustaqil")


@dp.message(OrderStates.entering_subject)
async def order_subject(message: types.Message, state: FSMContext):
    subject = message.text.strip()
    if len(subject) < 2 or len(subject) > 100:
        return await message.answer("❌ Fan nomi 2-100 ta belgi bo'lishi kerak.")
    await state.update_data(subject=subject)
    await state.set_state(OrderStates.entering_topic)
    await message.answer(
        f"✅ Fan: <b>{subject}</b>\n\n📝 Mavzuni kiriting:\n<i>Masalan: Ikkinchi jahon urushi sabablari</i>",
        parse_mode="HTML"
    )


@dp.message(OrderStates.entering_topic)
async def order_topic(message: types.Message, state: FSMContext):
    topic = message.text.strip()
    if len(topic) < 3 or len(topic) > 200:
        return await message.answer("❌ Mavzu 3-200 ta belgi bo'lishi kerak.")
    await state.update_data(topic=topic)
    await state.set_state(OrderStates.choosing_pages)
    await message.answer(
        f"✅ Mavzu: <b>{topic}</b>\n\n📄 Necha sahifa kerak? (maksimal 25)",
        reply_markup=kb_pages(), parse_mode="HTML"
    )


@dp.message(OrderStates.choosing_pages)
async def order_pages(message: types.Message, state: FSMContext):
    text = message.text.strip()
    if not text.isdigit() or not (1 <= int(text) <= 25):
        return await message.answer("❌ 1-25 orasida son kiriting:")
    await state.update_data(pages=int(text))
    await state.set_state(OrderStates.choosing_filetype)
    await message.answer(
        f"✅ Sahifalar: <b>{text}</b>\n\n📁 Fayl turini tanlang:",
        reply_markup=kb_filetype(), parse_mode="HTML"
    )


@dp.message(OrderStates.choosing_filetype)
async def order_filetype(message: types.Message, state: FSMContext):
    filetype = message.text.strip()
    if filetype not in FILETYPE_LABELS:
        return await message.answer("❌ Tugmalardan birini tanlang:")
    await state.update_data(filetype=filetype)
    await state.set_state(OrderStates.choosing_deadline)
    await message.answer(
        f"✅ Fayl turi: <b>{filetype}</b>\n\n⏰ Qachon tayyor bo'lishi kerak?",
        reply_markup=kb_deadline(), parse_mode="HTML"
    )


@dp.message(OrderStates.choosing_deadline)
async def order_deadline(message: types.Message, state: FSMContext):
    deadline = message.text.strip()
    if deadline not in DEADLINE_OPTIONS:
        return await message.answer("❌ Tugmalardan birini tanlang:")
    await state.update_data(deadline=deadline)
    await state.set_state(OrderStates.confirming)
    data          = await state.get_data()
    type_name     = ORDER_TYPES.get(data["order_type"], data["order_type"])
    deadline_date = datetime.now() + timedelta(days=int(deadline.split()[0]))
    summary = (
        f"📋 BUYURTMA MA'LUMOTLARI\n{'─'*25}\n"
        f"📌 Tur:       {type_name}\n📚 Fan:       {data['subject']}\n"
        f"📝 Mavzu:     {data['topic']}\n📄 Sahifa:    {data['pages']}\n"
        f"📁 Fayl turi: {data['filetype']}\n⏰ Muddat:    {deadline_date.strftime('%d.%m.%Y')} ({deadline})\n"
        f"{'─'*25}\n\n✅ Tasdiqlaysizmi?"
    )
    await message.answer(summary, reply_markup=kb_confirm())


@dp.message(OrderStates.confirming, F.text == "✅ Tasdiqlash")
async def order_confirm(message: types.Message, state: FSMContext):
    uid           = message.from_user.id
    data          = await state.get_data()
    deadline_str  = data["deadline"]
    deadline_date = deadline_str_to_date(deadline_str)
    filetype      = data.get("filetype", "📄 PDF")

    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute(
                    "INSERT INTO orders (user_id, type, subject, topic, pages, filetype, deadline, status, created_at) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,'pending',%s)",
                    (uid, data["order_type"], data["subject"], data["topic"], data["pages"], filetype, deadline_date, datetime.now())
                )
                order_id = c.lastrowid
        finally:
            db.close()
    except Exception as e:
        print(f"[ORDER INSERT ERROR] {e}")
        return await message.answer("❌ Buyurtmani saqlashda xato.")

    await state.clear()
    type_name = ORDER_TYPES.get(data["order_type"], data["order_type"])
    await message.answer(
        f"✅ Buyurtma qabul qilindi!\n🆔 #{order_id}\n📁 {filetype}\n⏰ {deadline_date}\n⏳ Admin ko'rib chiqadi.",
        reply_markup=get_menu(uid)
    )
    username  = message.from_user.username or "—"
    admin_txt = (
        f"🆕 <b>YANGI BUYURTMA #{order_id}</b>\n{'─'*25}\n"
        f"👤 @{username} (ID: <code>{uid}</code>)\n"
        f"📌 {type_name}\n📚 {data['subject']}\n📝 {data['topic']}\n"
        f"📄 {data['pages']} sahifa\n📁 {filetype}\n⏰ {deadline_date} ({deadline_str})\n"
        f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    admin_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Qabul qilish", callback_data=f"oAccept:{order_id}:{uid}"),
        InlineKeyboardButton(text="❌ Rad etish",    callback_data=f"oReject:{order_id}:{uid}")
    ]])
    targets = [OWNER_ID] + [a["user_id"] for a in all_sub_admins() if "orders" in json.loads(a.get("perms", "[]"))]
    for t in targets:
        try:
            await bot.send_message(t, admin_txt, reply_markup=admin_kb, parse_mode="HTML")
        except Exception as e:
            print(f"[ORDER NOTIFY ERROR uid={t}] {e}")

# ================= ADMIN ORDER CALLBACKS =================
@dp.callback_query(F.data.startswith("oAccept:"))
async def order_accept(callback: types.CallbackQuery):
    if not has_perm(callback.from_user.id, "orders"):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    try:
        _, order_id, uid = callback.data.split(":")
        order_id, uid = int(order_id), int(uid)
    except Exception:
        return await callback.answer("❌ Noto'g'ri format", show_alert=True)
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("UPDATE orders SET status='in_progress' WHERE id=%s", (order_id,))
        finally:
            db.close()
    except Exception as e:
        print(f"[ORDER ACCEPT ERROR] {e}")
        return await callback.answer("❌ DB xato", show_alert=True)
    try:
        await bot.send_message(uid, f"✅ Buyurtmangiz #{order_id} qabul qilindi!\n⏳ Tayyor bo'lganda yuboriladi.")
    except Exception as e:
        print(f"[SEND ERROR uid={uid}] {e}")
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.message.reply(
            f"✅ #{order_id} qabul qilindi.\n\n📎 Fayl: <code>/deliver {order_id}</code>\n✏️ Matn: <code>/delivertext {order_id}</code>",
            parse_mode="HTML"
        )
    except Exception:
        pass
    await callback.answer("✅ Qabul qilindi")


@dp.callback_query(F.data.startswith("oReject:"))
async def order_reject_cb(callback: types.CallbackQuery):
    if not has_perm(callback.from_user.id, "orders"):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    try:
        _, order_id, uid = callback.data.split(":")
        order_id, uid = int(order_id), int(uid)
    except Exception:
        return await callback.answer("❌ Noto'g'ri format", show_alert=True)
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("UPDATE orders SET status='rejected' WHERE id=%s", (order_id,))
        finally:
            db.close()
    except Exception as e:
        print(f"[ORDER REJECT ERROR] {e}")
        return await callback.answer("❌ DB xato", show_alert=True)
    try:
        await bot.send_message(uid, f"❌ Buyurtmangiz #{order_id} rad etildi.")
    except Exception as e:
        print(f"[SEND ERROR uid={uid}] {e}")
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.message.reply(f"❌ #{order_id} rad etildi.")
    except Exception:
        pass
    await callback.answer("❌ Rad etildi")

# ================= DELIVER — FILE =================
@dp.message(Command("deliver"))
async def deliver_file_cmd(message: types.Message, state: FSMContext):
    if not has_perm(message.from_user.id, "orders"):
        return
    args = message.text.split()
    if len(args) < 2 or not args[1].isdigit():
        return await message.answer("❌ Ishlatish: /deliver {order_id}")
    order_id = int(args[1])
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT id, user_id, type, subject, topic, status FROM orders WHERE id=%s", (order_id,))
                order = c.fetchone()
        finally:
            db.close()
    except Exception:
        return await message.answer("❌ DB xato")
    if not order:
        return await message.answer(f"❌ #{order_id} topilmadi")
    if order["status"] == "done":
        return await message.answer(f"⚠️ #{order_id} allaqachon bajarilgan")
    type_name = ORDER_TYPES.get(order["type"], order["type"])
    await state.set_state(DeliverStates.waiting_file)
    await state.update_data(order_id=order_id, target_uid=order["user_id"])
    await message.answer(
        f"📤 #{order_id} uchun fayl yuboring:\n👤 {order['user_id']}\n📌 {type_name} | {order['subject']} — {order['topic']}\n\nFayl yuboring:",
        reply_markup=kb_cancel()
    )


@dp.message(DeliverStates.waiting_file, F.document)
async def deliver_file_send(message: types.Message, state: FSMContext):
    data       = await state.get_data()
    order_id   = data["order_id"]
    target_uid = data["target_uid"]
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT subject, topic, type FROM orders WHERE id=%s", (order_id,))
                order = c.fetchone()
            with db.cursor() as c:
                c.execute("UPDATE orders SET status='done' WHERE id=%s", (order_id,))
        finally:
            db.close()
    except Exception:
        await message.answer("❌ DB xato")
        await state.clear()
        return
    type_name = ORDER_TYPES.get(order["type"], order["type"])
    caption   = f"✅ Buyurtmangiz tayyor!\n🆔 #{order_id} | {type_name}\n📚 {order['subject']} — {order['topic']}"
    try:
        await bot.send_document(target_uid, message.document.file_id, caption=caption)
        await message.answer(f"✅ Fayl yuborildi (#{order_id})", reply_markup=get_menu(message.from_user.id))
    except Exception as e:
        await message.answer(f"❌ Xato: {e}")
    await state.clear()

# ================= DELIVER — TEXT =================
@dp.message(Command("delivertext"))
async def deliver_text_cmd(message: types.Message, state: FSMContext):
    if not has_perm(message.from_user.id, "orders"):
        return
    args = message.text.split()
    if len(args) < 2 or not args[1].isdigit():
        return await message.answer("❌ Ishlatish: /delivertext {order_id}")
    order_id = int(args[1])
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT id, user_id, type, subject, topic, status FROM orders WHERE id=%s", (order_id,))
                order = c.fetchone()
        finally:
            db.close()
    except Exception:
        return await message.answer("❌ DB xato")
    if not order:
        return await message.answer(f"❌ #{order_id} topilmadi")
    if order["status"] == "done":
        return await message.answer(f"⚠️ #{order_id} allaqachon bajarilgan")
    type_name = ORDER_TYPES.get(order["type"], order["type"])
    await state.set_state(DeliverStates.waiting_text)
    await state.update_data(order_id=order_id, target_uid=order["user_id"])
    await message.answer(
        f"✏️ #{order_id} uchun matn yuboring:\n👤 {order['user_id']}\n📌 {type_name} | {order['subject']} — {order['topic']}\n\nMatnni yozing:",
        reply_markup=kb_cancel()
    )


@dp.message(DeliverStates.waiting_text, F.text)
async def deliver_text_send(message: types.Message, state: FSMContext):
    data       = await state.get_data()
    order_id   = data["order_id"]
    target_uid = data["target_uid"]
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT subject, topic, type FROM orders WHERE id=%s", (order_id,))
                order = c.fetchone()
            with db.cursor() as c:
                c.execute("UPDATE orders SET status='done' WHERE id=%s", (order_id,))
        finally:
            db.close()
    except Exception:
        await message.answer("❌ DB xato")
        await state.clear()
        return
    type_name = ORDER_TYPES.get(order["type"], order["type"])
    header    = f"✅ Buyurtmangiz tayyor!\n🆔 #{order_id} | {type_name}\n📚 {order['subject']} — {order['topic']}\n{'─'*25}\n\n"
    try:
        await bot.send_message(target_uid, header + message.text)
        await message.answer(f"✅ Matn yuborildi (#{order_id})", reply_markup=get_menu(message.from_user.id))
    except Exception as e:
        await message.answer(f"❌ Xato: {e}")
    await state.clear()

# ================= MY ORDERS =================
@dp.message(F.text == "📋 Buyurtmalarim")
async def my_orders(message: types.Message):
    uid = message.from_user.id
    if not has_access(uid):
        return await message.answer("❌ Bu funksiya premium foydalanuvchilar uchun.")
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute(
                    "SELECT id, type, subject, topic, pages, deadline, status, created_at "
                    "FROM orders WHERE user_id=%s ORDER BY id DESC LIMIT 10",
                    (uid,)
                )
                rows = c.fetchall()
        finally:
            db.close()
    except Exception:
        return await message.answer("❌ Xato yuz berdi.")
    if not rows:
        return await message.answer("📋 Hali buyurtmalar yo'q.")
    used      = get_order_monthly_count(uid)
    remaining = max(0, ORDER_LIMIT_PER_MONTH - used)
    status_icons = {
        "pending": "⏳ Kutilmoqda", "in_progress": "🔄 Jarayonda",
        "done": "✅ Bajarildi", "rejected": "❌ Rad etildi"
    }
    txt = f"📋 BUYURTMALARIM\n📊 Bu oy: {used}/{ORDER_LIMIT_PER_MONTH} (qoldi: {remaining})\n{'─'*25}\n\n"
    for r in rows:
        txt += (
            f"🆔 #{r['id']} | {ORDER_TYPES.get(r['type'], r['type'])}\n"
            f"📚 {r['subject']} — {r['topic']}\n"
            f"📄 {r['pages']} sahifa | ⏰ {format_deadline(r['deadline'])}\n"
            f"📅 {r['created_at'].strftime('%d.%m.%Y') if r['created_at'] else '—'} | {status_icons.get(r['status'], r['status'])}\n{'─'*25}\n"
        )
    if len(txt) > 4000:
        txt = txt[:4000] + "\n..."
    await message.answer(txt)

# ================================================================
# ═════════════════  BROADCAST  ══════════════════════════════════
# ================================================================

@dp.message(Command("broadcast"))
@dp.callback_query(F.data == "adm:broadcast")
async def broadcast_cmd(event: types.Message | types.CallbackQuery, state: FSMContext):
    uid = event.from_user.id
    if not has_perm(uid, "broadcast"):
        if isinstance(event, types.CallbackQuery):
            return await event.answer("❌ Ruxsat yo'q", show_alert=True)
        return
    await state.set_state(BroadcastStates.waiting_message)
    text = "📢 <b>BROADCAST</b>\n\nXabar yuboring — har qanday turdagi:\n• Matn\n• Rasm\n• Fayl\n• Video\n• Forward"
    if isinstance(event, types.CallbackQuery):
        await event.answer()
        await event.message.answer(text, reply_markup=kb_cancel(), parse_mode="HTML")
    else:
        await event.answer(text, reply_markup=kb_cancel(), parse_mode="HTML")


@dp.message(BroadcastStates.waiting_message)
async def broadcast_preview(message: types.Message, state: FSMContext):
    if message.forward_from or message.forward_from_chat:
        msg_type = "forward"
    elif message.photo:    msg_type = "photo"
    elif message.video:    msg_type = "video"
    elif message.document: msg_type = "document"
    elif message.audio:    msg_type = "audio"
    elif message.voice:    msg_type = "voice"
    elif message.sticker:  msg_type = "sticker"
    elif message.text:     msg_type = "text"
    else:
        return await message.answer("❌ Bu turdagi xabar qo'llab-quvvatlanmaydi.")
    await state.update_data(
        msg_type=msg_type,
        from_chat_id=message.chat.id,
        message_id=message.message_id,
        original_text=message.text or message.caption or ""
    )
    user_count = len(get_all_user_ids())
    await state.set_state(BroadcastStates.confirming)
    await message.answer(
        f"📢 <b>BROADCAST PREVIEW</b>\n\n👥 Yuboriladi: <b>{user_count} ta</b>\n📨 Tur: <b>{msg_type}</b>\n\nTasdiqlaysizmi?",
        reply_markup=kb_broadcast_confirm(), parse_mode="HTML"
    )


@dp.callback_query(F.data == "bc:send")
async def broadcast_send(callback: types.CallbackQuery, state: FSMContext):
    if not has_perm(callback.from_user.id, "broadcast"):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    data         = await state.get_data()
    user_ids     = get_all_user_ids()
    msg_type     = data.get("msg_type", "")
    original_txt = data.get("original_text", "")
    await state.clear()
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("⏳ Yuborilmoqda...")
    status_msg = await callback.message.reply(f"⏳ Yuborilmoqda... 0/{len(user_ids)}")
    success = failed = 0
    for i, uid in enumerate(user_ids, 1):
        try:
            if msg_type == "text" and original_txt:
                await bot.send_message(
                    chat_id=uid,
                    text=f"<b>Admin:</b>\n\n{original_txt}",
                    parse_mode="HTML"
                )
            else:
                await bot.copy_message(
                    chat_id=uid,
                    from_chat_id=data["from_chat_id"],
                    message_id=data["message_id"]
                )
            success += 1
        except Exception:
            failed += 1
        if i % 20 == 0:
            try:
                await status_msg.edit_text(f"⏳ {i}/{len(user_ids)}\n✅ {success} | ❌ {failed}")
            except Exception:
                pass
        await asyncio.sleep(0.05)
    try:
        await status_msg.edit_text(
            f"✅ <b>Broadcast yakunlandi!</b>\n\n👥 Jami: {len(user_ids)}\n✅ {success} | ❌ {failed}",
            parse_mode="HTML"
        )
    except Exception:
        pass


@dp.callback_query(F.data == "bc:cancel")
async def broadcast_cancel(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("❌ Bekor qilindi")
    await callback.message.reply("❌ Broadcast bekor qilindi.")

# ================================================================
# ═════════════════  REVOKE PREMIUM  ═════════════════════════════
# ================================================================

@dp.message(Command("revoke"))
@dp.callback_query(F.data == "adm:revoke")
async def revoke_cmd(event: types.Message | types.CallbackQuery, state: FSMContext):
    uid = event.from_user.id
    if not has_perm(uid, "revoke"):
        if isinstance(event, types.CallbackQuery):
            return await event.answer("❌ Ruxsat yo'q", show_alert=True)
        return
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute(
                    "SELECT user_id, username, access_until FROM users WHERE access_until > %s ORDER BY access_until DESC",
                    (datetime.now(),)
                )
                rows = c.fetchall()
        finally:
            db.close()
    except Exception:
        rows = []
    if not rows:
        text = "❌ Hozirda premium foydalanuvchilar yo'q."
        if isinstance(event, types.CallbackQuery):
            await event.answer()
            await event.message.answer(text)
        else:
            await event.answer(text)
        return
    buttons = []
    for r in rows:
        uname        = r["username"] or "—"
        access_until = r["access_until"].strftime('%d.%m.%Y') if r["access_until"] else "—"
        buttons.append([InlineKeyboardButton(text=f"🚫 @{uname} | {access_until}", callback_data=f"revokeUser:{r['user_id']}")])
    buttons.append([InlineKeyboardButton(text="❌ Yopish", callback_data="revokeClose")])
    kb   = InlineKeyboardMarkup(inline_keyboard=buttons)
    text = f"🚫 <b>PREMIUM BEKOR QILISH</b>\n\nUserni tanlang:\n(Jami: {len(rows)} ta premium user)"
    if isinstance(event, types.CallbackQuery):
        await event.answer()
        await event.message.answer(text, reply_markup=kb, parse_mode="HTML")
    else:
        await event.answer(text, reply_markup=kb, parse_mode="HTML")


@dp.callback_query(F.data.startswith("revokeUser:"))
async def revoke_user(callback: types.CallbackQuery):
    if not has_perm(callback.from_user.id, "revoke"):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    try:
        uid = int(callback.data.split(":")[1])
    except Exception:
        return await callback.answer("❌ Noto'g'ri format", show_alert=True)
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT username, access_until FROM users WHERE user_id=%s", (uid,))
                user = c.fetchone()
            if not user:
                return await callback.answer("❌ User topilmadi", show_alert=True)
            with db.cursor() as c:
                c.execute("UPDATE users SET access_until=NULL, warned=0 WHERE user_id=%s", (uid,))
        finally:
            db.close()
    except Exception:
        return await callback.answer("❌ DB xato", show_alert=True)
    try:
        await bot.send_message(uid, "⚠️ Sizning premium obunangiz bekor qilindi.", reply_markup=menu_basic)
    except Exception:
        pass
    uname = user["username"] or "—"
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.message.reply(f"✅ @{uname} (ID: {uid}) premiumdan chiqarildi.")
    except Exception:
        pass
    await callback.answer(f"✅ @{uname} premiumdan chiqarildi")


@dp.callback_query(F.data == "revokeClose")
async def revoke_close(callback: types.CallbackQuery):
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("✅ Yopildi")

# ================================================================
# ═════════════════  ADMIN PANEL  ════════════════════════════════
# ================================================================

@dp.message(Command("admin"))
async def admin_panel(message: types.Message):
    uid = message.from_user.id
    if not is_any_admin(uid):
        return
    stats = get_admin_stats()
    await message.answer(build_admin_text(stats, uid), reply_markup=admin_panel_kb(uid), parse_mode="HTML")


@dp.callback_query(F.data == "adm:refresh")
async def adm_refresh(callback: types.CallbackQuery):
    uid = callback.from_user.id
    if not is_any_admin(uid):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    stats = get_admin_stats()
    try:
        await callback.message.edit_text(build_admin_text(stats, uid), reply_markup=admin_panel_kb(uid), parse_mode="HTML")
    except Exception:
        pass
    await callback.answer("🔄 Yangilandi")


@dp.callback_query(F.data == "adm:users")
async def adm_users(callback: types.CallbackQuery):
    if not has_perm(callback.from_user.id, "users"):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute(
                    "SELECT user_id, username, access_until FROM users WHERE access_until > %s ORDER BY access_until DESC",
                    (datetime.now(),)
                )
                rows = c.fetchall()
        finally:
            db.close()
    except Exception:
        rows = []
    await callback.answer()
    if not rows:
        return await callback.message.answer("❌ Premium userlar yo'q")
    txt = "✅ <b>PREMIUM USERLAR:</b>\n\n"
    for r in rows:
        username     = r["username"] or "—"
        access_until = r["access_until"].strftime('%d.%m.%Y') if r["access_until"] else "—"
        txt += f"@{username} (<code>{r['user_id']}</code>) | {access_until}\n"
    if len(txt) > 4000:
        txt = txt[:4000] + "\n..."
    await callback.message.answer(txt, parse_mode="HTML")


@dp.callback_query(F.data == "adm:payments")
async def adm_payments(callback: types.CallbackQuery):
    if not has_perm(callback.from_user.id, "payments"):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT id, user_id, tariff_days, status, created_at FROM payments ORDER BY id DESC LIMIT 10")
                rows = c.fetchall()
        finally:
            db.close()
    except Exception:
        rows = []
    await callback.answer()
    if not rows:
        return await callback.message.answer("❌ To'lovlar yo'q")
    txt = "💳 <b>OXIRGI TO'LOVLAR:</b>\n\n"
    for r in rows:
        icon     = "✅" if r["status"] == "approved" else ("❌" if r["status"] == "rejected" else "⏳")
        date_str = r["created_at"].strftime('%d.%m.%Y') if r["created_at"] else "—"
        txt += f"{icon} #{r['id']} | <code>{r['user_id']}</code> | {r['tariff_days'] or '?'} kun | {date_str}\n"
    await callback.message.answer(txt, parse_mode="HTML")


@dp.callback_query(F.data == "adm:expiring")
async def adm_expiring(callback: types.CallbackQuery):
    if not has_perm(callback.from_user.id, "orders"):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    now  = datetime.now()
    soon = now + timedelta(days=3)
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT user_id, username, access_until FROM users WHERE access_until BETWEEN %s AND %s", (now, soon))
                rows = c.fetchall()
        finally:
            db.close()
    except Exception:
        rows = []
    await callback.answer()
    if not rows:
        return await callback.message.answer("✅ Tugayotgan obunalar yo'q")
    txt = "⏰ <b>TUGAYOTGAN OBUNALAR (3 kun):</b>\n\n"
    for r in rows:
        username     = r["username"] or "—"
        access_until = r["access_until"].strftime('%d.%m.%Y %H:%M') if r["access_until"] else "—"
        txt += f"@{username} (<code>{r['user_id']}</code>) | {access_until}\n"
    await callback.message.answer(txt, parse_mode="HTML")


@dp.callback_query(F.data == "adm:orders")
async def adm_orders(callback: types.CallbackQuery):
    if not has_perm(callback.from_user.id, "orders"):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute(
                    "SELECT o.id, o.user_id, u.username, o.type, o.subject, o.topic, o.pages, o.deadline, o.status, o.created_at "
                    "FROM orders o LEFT JOIN users u ON o.user_id = u.user_id "
                    "WHERE o.status IN ('pending','in_progress') ORDER BY o.id DESC"
                )
                rows = c.fetchall()
        finally:
            db.close()
    except Exception:
        rows = []
    await callback.answer()
    if not rows:
        return await callback.message.answer("✅ Aktiv buyurtmalar yo'q.")
    status_icons = {"pending": "⏳", "in_progress": "🔄"}
    txt = f"📋 <b>AKTIV BUYURTMALAR ({len(rows)} ta)</b>\n{'─'*25}\n\n"
    for r in rows:
        username     = r["username"] or "—"
        type_name    = ORDER_TYPES.get(r["type"], r["type"])
        icon         = status_icons.get(r["status"], "")
        date_str     = r["created_at"].strftime('%d.%m.%Y') if r["created_at"] else "—"
        deadline_txt = format_deadline(r["deadline"])
        txt += (
            f"{icon} #{r['id']} | {type_name}\n👤 @{username} (<code>{r['user_id']}</code>)\n"
            f"📚 {r['subject']} — {r['topic']}\n📄 {r['pages']} s. | ⏰ {deadline_txt} | 📅 {date_str}\n"
            f"/deliver {r['id']} | /delivertext {r['id']}\n{'─'*25}\n"
        )
    if len(txt) > 4000:
        txt = txt[:4000] + "\n..."
    await callback.message.answer(txt, parse_mode="HTML")

# ================= SLASH COMMANDS =================
@dp.message(Command("users"))
async def users_cmd(message: types.Message):
    if not has_perm(message.from_user.id, "users"):
        return
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT user_id, username, access_until FROM users WHERE access_until > %s ORDER BY access_until DESC", (datetime.now(),))
                rows = c.fetchall()
        finally:
            db.close()
    except Exception:
        return await message.answer("❌ DB xato")
    if not rows:
        return await message.answer("❌ Premium userlar yo'q")
    txt = "✅ <b>PREMIUM USERLAR:</b>\n\n"
    for r in rows:
        username     = r["username"] or "—"
        access_until = r["access_until"].strftime('%d.%m.%Y') if r["access_until"] else "—"
        txt += f"@{username} (<code>{r['user_id']}</code>) | {access_until}\n"
    if len(txt) > 4000:
        txt = txt[:4000] + "\n..."
    await message.answer(txt, parse_mode="HTML")


@dp.message(Command("payments"))
async def payments_cmd(message: types.Message):
    if not has_perm(message.from_user.id, "payments"):
        return
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT id, user_id, tariff_days, status, created_at FROM payments ORDER BY id DESC LIMIT 10")
                rows = c.fetchall()
        finally:
            db.close()
    except Exception:
        return await message.answer("❌ DB xato")
    if not rows:
        return await message.answer("❌ To'lovlar yo'q")
    txt = "💳 <b>OXIRGI TO'LOVLAR:</b>\n\n"
    for r in rows:
        icon     = "✅" if r["status"] == "approved" else ("❌" if r["status"] == "rejected" else "⏳")
        date_str = r["created_at"].strftime('%d.%m.%Y') if r["created_at"] else "—"
        txt += f"{icon} #{r['id']} | <code>{r['user_id']}</code> | {r['tariff_days'] or '?'} kun | {date_str}\n"
    await message.answer(txt, parse_mode="HTML")


@dp.message(Command("expiring"))
async def expiring_cmd(message: types.Message):
    if not has_perm(message.from_user.id, "orders"):
        return
    now  = datetime.now()
    soon = now + timedelta(days=3)
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute("SELECT user_id, username, access_until FROM users WHERE access_until BETWEEN %s AND %s", (now, soon))
                rows = c.fetchall()
        finally:
            db.close()
    except Exception:
        return await message.answer("❌ DB xato")
    if not rows:
        return await message.answer("✅ Tugayotgan obunalar yo'q")
    txt = "⏰ <b>TUGAYOTGAN OBUNALAR (3 kun):</b>\n\n"
    for r in rows:
        username     = r["username"] or "—"
        access_until = r["access_until"].strftime('%d.%m.%Y %H:%M') if r["access_until"] else "—"
        txt += f"@{username} (<code>{r['user_id']}</code>) | {access_until}\n"
    await message.answer(txt, parse_mode="HTML")


@dp.message(Command("orders"))
async def orders_cmd(message: types.Message):
    if not has_perm(message.from_user.id, "orders"):
        return
    try:
        db = get_db()
        try:
            with db.cursor() as c:
                c.execute(
                    "SELECT o.id, o.user_id, u.username, o.type, o.subject, o.topic, o.pages, o.deadline, o.status, o.created_at "
                    "FROM orders o LEFT JOIN users u ON o.user_id = u.user_id "
                    "WHERE o.status IN ('pending','in_progress') ORDER BY o.id DESC"
                )
                rows = c.fetchall()
        finally:
            db.close()
    except Exception:
        return await message.answer("❌ DB xato")
    if not rows:
        return await message.answer("✅ Aktiv buyurtmalar yo'q.")
    txt = f"📋 <b>AKTIV BUYURTMALAR ({len(rows)} ta)</b>\n{'─'*25}\n\n"
    for r in rows:
        username     = r["username"] or "—"
        type_name    = ORDER_TYPES.get(r["type"], r["type"])
        date_str     = r["created_at"].strftime('%d.%m.%Y') if r["created_at"] else "—"
        deadline_txt = format_deadline(r["deadline"])
        txt += (
            f"🆔 #{r['id']} | {type_name}\n👤 @{username} (<code>{r['user_id']}</code>)\n"
            f"📚 {r['subject']} — {r['topic']}\n📄 {r['pages']} s. | ⏰ {deadline_txt} | 📅 {date_str}\n"
            f"/deliver {r['id']} | /delivertext {r['id']}\n{'─'*25}\n"
        )
    if len(txt) > 4000:
        txt = txt[:4000] + "\n..."
    await message.answer(txt, parse_mode="HTML")

# ================================================================
# ═════════════════  WATCHER  ════════════════════════════════════
# ================================================================

async def access_watcher():
    while True:
        try:
            db = get_db()
            try:
                with db.cursor() as c:
                    c.execute("SELECT user_id, access_until FROM users WHERE warned=0 AND access_until IS NOT NULL AND access_until > NOW()")
                    rows = c.fetchall()
            finally:
                db.close()
            for r in rows:
                delta = r["access_until"] - datetime.now()
                if timedelta(days=0) < delta <= timedelta(days=3):
                    try:
                        await bot.send_message(r["user_id"], "⏰ Diqqat! Obunangiz tugashiga 3 kun qoldi.\nUzluksiz foydalanish uchun to'lovni yangilang.")
                        db2 = get_db()
                        try:
                            with db2.cursor() as c2:
                                c2.execute("UPDATE users SET warned=1 WHERE user_id=%s", (r["user_id"],))
                        finally:
                            db2.close()
                    except Exception:
                        pass
        except Exception as e:
            print(f"[Watcher ERROR] {e}")
        await asyncio.sleep(3600)

# ================================================================
# ═════════════════  STARTUP & RUN  ══════════════════════════════
# ================================================================

async def on_startup():
    ensure_admins_table()
    asyncio.create_task(access_watcher())
    print("✅ Bot ishga tushdi")
    if HAS_CV2:
        print("✅ OpenCV mavjud — auto-crop va professional skan yoqildi")
    else:
        print("⚠️  OpenCV yo'q — fallback skan rejimi ishlatiladi (pip install opencv-python-headless)")


async def main():
    print("🚀 Starting bot...")
    dp.startup.register(on_startup)
    while True:
        try:
            await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types(), close_bot_session=False)
        except Exception as e:
            print(f"[POLLING ERROR] {e} — restarting in 5s...")
            await asyncio.sleep(5)


if __name__ == "__main__":
    asyncio.run(main())